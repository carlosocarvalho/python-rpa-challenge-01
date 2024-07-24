

from selenium import webdriver # type: ignore

from selenium.webdriver.common.keys import Keys # type: ignore
from selenium.webdriver.common.by import By # type: ignore
from time import sleep # type: ignore
from openpyxl import load_workbook # type: ignore

book =  load_workbook('challenge.xlsx')

sheet = book['Sheet1']

browser = webdriver.Chrome()
browser.get('https://rpachallenge.com/')

# input('Press Enter to continue...')

sleep(.9)

#start filling out the form

browser.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button').click()

# interact with the form
for row in sheet.iter_rows(min_row=2):
     if row[0].value == None:
        break
     firstName = row[0].value
     lastName = row[1].value
     conpanyName = row[2].value
     roleInCompany = row[3].value
     address = row[4].value
     email = row[5].value
     phoneNumber = row[6].value
   
     # populate the form  
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelFirstName"]').send_keys(firstName)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelLastName"]').send_keys(lastName)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelCompanyName"]').send_keys(conpanyName)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelRole"]').send_keys(roleInCompany)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelAddress"]').send_keys(address)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelEmail"]').send_keys(email)
     browser.find_element(By.XPATH, '//input[@ng-reflect-name="labelPhone"]').send_keys(phoneNumber)
     browser.find_element(By.XPATH, '//input[@type="submit"]').click()
     sleep(2)


input('Press Enter to continue...')
# browser.close()